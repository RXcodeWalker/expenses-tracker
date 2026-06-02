from openpyxl import Workbook, load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
import os


FILE_NAME = "expenses.xlsx"


class ExpenseTracker:
    def __init__(self):
        self.setup_file()

    def setup_file(self):
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"

            ws.append([
                "Date",
                "Description",
                "Category",
                "Amount"
            ])

            wb.save(FILE_NAME)

    def add_expense(self):
        description = input("Description: ")
        category = input("Category: ")

        try:
            amount = float(input("Amount: ₹"))
        except ValueError:
            print("Invalid amount.")
            return

        wb = load_workbook(FILE_NAME)
        ws = wb["Expenses"]

        ws.append([
            datetime.now().strftime("%Y-%m-%d"),
            description,
            category,
            amount
        ])

        wb.save(FILE_NAME)

        print("Expense added successfully.")

    def view_expenses(self):
        wb = load_workbook(FILE_NAME)
        ws = wb["Expenses"]

        print("\n--- Expenses ---")

        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            count += 1

            print(
                f"{count}. "
                f"{row[0]} | "
                f"{row[1]} | "
                f"{row[2]} | "
                f"₹{row[3]}"
            )

        if count == 0:
            print("No expenses found.")

    def total_spending(self):
        wb = load_workbook(FILE_NAME)
        ws = wb["Expenses"]

        total = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            total += row[3]

        print(f"\nTotal Spending: ₹{total:.2f}")

    def category_summary(self):
        wb = load_workbook(FILE_NAME)
        ws = wb["Expenses"]

        categories = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            category = row[2]
            amount = row[3]

            categories[category] = (
                categories.get(category, 0)
                + amount
            )

        print("\n--- Category Summary ---")

        for category, total in categories.items():
            print(f"{category}: ₹{total:.2f}")

    def generate_chart(self):
        wb = load_workbook(FILE_NAME)
        ws = wb["Expenses"]

        categories = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            category = row[2]
            amount = row[3]

            categories[category] = (
                categories.get(category, 0)
                + amount
            )

        if not categories:
            print("No data available.")
            return

        plt.pie(
            categories.values(),
            labels=categories.keys(),
            autopct="%1.1f%%"
        )

        plt.title("Expense Breakdown")
        plt.show()

    def menu(self):
        while True:

            print("\n======Expense Tracker =====")
            print("1. Add Expense")
            print("2. View Expenses")
            print("3. Total Spending")
            print("4. Category Summary")
            print("5. Generate Pie Chart")
            print("6. Exit")

            choice = input("\nEnter choice: ")

            if choice == "1":
                self.add_expense()

            elif choice == "2":
                self.view_expenses()

            elif choice == "3":
                self.total_spending()

            elif choice == "4":
                self.category_summary()

            elif choice == "5":
                self.generate_chart()

            elif choice == "6":
                print("Goodbye.")
                break

            else:
                print("Invalid choice.")


if __name__ == "__main__":
    tracker = ExpenseTracker()
    tracker.menu()